import logging
from markupsafe import Markup
from openpyxl.utils import get_column_letter

from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from copy import copy
import random

_logger = logging.getLogger(__name__)


class OutboundOrder(models.Model):
    _name = 'world.depot.outbound.order'
    _description = 'Outbound Order'
    _inherit = ["mail.thread", "mail.activity.mixin"]
    _rec_name = 'billno'

    # Fields
    type = fields.Selection(
        selection=[
            ('outbound', 'Outbound'),
            ('service', 'Service'),
            ('transfer', 'Transfer'),
        ],
        default='outbound',
        string="Order Type",
        required=True,
        tracking=True
    )
    billno = fields.Char(string='BillNo', readonly=True)
    date = fields.Date(string='Order Date', required=True, tracking=True, default=fields.Date.today)
    p_date = fields.Date(string='Planning Date', required=False, tracking=True)
    o_date = fields.Date(string='Outbound Date', required=False, tracking=True)
    project = fields.Many2one('project.project', string='Project', required=True)
    owner = fields.Many2one('res.partner', string='Owner', related='project.owner', stored=True, tracking=True)
    warehouse = fields.Many2one('stock.warehouse', string='Warehouse', tracking=True,
                                stored=True)
    pick_type = fields.Many2one('stock.picking.type', string='Picking Type',
                                tracking=True, help='Picking type for this project', stored=True,
                                domain=[('code', '=', 'internal')])
    unload_company = fields.Many2one('res.partner', string='Unload Company/Person', required=True, tracking=True)
    remark = fields.Text(string='Remark')
    remark1 = fields.Text(string='Remark 1')
    reference = fields.Char(string='Reference', required=True, help='Reference for the Order No of Owner', )
    load_ref = fields.Char(string='Loading Reference', required=False, help='Reference for the Delivery', )
    load_date = fields.Datetime(string='Loading Date', required=False, tracking=True,
                                help='Date when the loading was completed')
    delivery_method = fields.Selection(
        selection=[
            ('truck', 'Truck Delivery'),
            ('pickup', 'Customer Pickup'),
            ('parcel', 'Parcel Delivery'),  # Modified from "express" to "parcel"
        ],
        string="Delivery Method",
        required=False,
        default='truck',
        tracking=True,
        help="Parcel Delivery refers to standard package shipping services"
    )
    delivery_company = fields.Many2one(
        comodel_name='res.partner',
        string='Delivery Company',
        required=False,
        tracking=True,
        help='Company providing the truck for delivery'
    )
    delivery_number = fields.Char(
        string='Delivery Number',
        required=False,
        tracking=True,
        help='Delivery number for the truck'
    )
    # Address Information
    delivery_street = fields.Char(string='Delivery Street', required=False, tracking=True)
    delivery_city = fields.Char(string='Delivery City', required=False, tracking=True)
    delivery_zip = fields.Char(string='Delivery Zip', required=False, tracking=True)
    delivery_country_id = fields.Many2one('res.country', string='Delivery Country', required=False, tracking=True)
    delivery_phone = fields.Char(string='Delivery Phone', required=False, tracking=True)
    delivery_mobile = fields.Char(string='Delivery Mobile', required=False, tracking=True)
    delivery_email = fields.Char(string='Delivery Email', required=False, tracking=True)

    # 收货人签回日期
    receiver_sign_back_date = fields.Datetime(string='Receiver Sign Back Date', readonly=False, tracking=True)
    # 司机签收日期
    driver_sign_date = fields.Datetime(string='Driver Sign Date', readonly=False, tracking=True)
    plate_number = fields.Char(string='Plate Number', required=False, tracking=True,
                               help='Vehicle plate number for the delivery truck')
    delivery_remark = fields.Text(string='Delivery Remark',
                                  help='Additional remarks for the delivery process')
    pod_file = fields.Binary(
        string='POD File',
        help='Proof of Delivery file, such as a signed document or image',
        attachment=True,
        tracking=True
    )
    pod_filename = fields.Char(
        string='POD Filename',
        help='Filename of the Proof of Delivery file',
        tracking=True
    )

    state = fields.Selection(
        selection=[
            ('new', 'New'),
            ('confirm', 'Confirm'),
            ('cancel', 'Cancel')
        ],
        default='new',
        string="State",
        tracking=True
    )
    outbound_order_product_ids = fields.One2many(
        comodel_name='world.depot.outbound.order.product',
        inverse_name='outbound_order_id',
        string='Products of Outbound Order',
    )
    outbound_order_docs_ids = fields.One2many(
        comodel_name='world.depot.outbound.order.docs',
        inverse_name='outbound_order_id',
        string='Other Documents',
        help='Other documents related to the outbound order, such as invoices, packing lists, etc.'
    )

    picking_PICK = fields.Many2one('stock.picking', string='Picking', readonly=True,
                                   help='Reference to the related Stock Picking')
    picking_PICK_date = fields.Datetime(string='Picking Date', readonly=True,
                                        help='Date when the stock picking was validated')
    picking_Out = fields.Many2one('stock.picking', string='Outbound', readonly=True,
                                  help='Reference to the related Stock Out')
    picking_Out_date = fields.Datetime(string='Outbound Date', readonly=True,
                                       help='Date when the stock delivery was validated')
    status = fields.Selection(
        selection=[
            ('planning', 'Planning'),
            ('picking', 'Picking'),
            ('outbound', 'Outbound'),
        ],
        default='planning',
        string="Status",
        readonly=True,
        tracking=True
    )
    pallets = fields.Float(string='Pallets')
    scanning_quantity = fields.Float(string='Scanning Quantity', default=0.0, tracking=True)
    is_adr = fields.Boolean(string='ADR', default=True, tracking=True)

    # 尾程配送 (Last Mile Delivery Charge)
    outbound_delivery_charge = fields.Float(
        string='Delivery Charge',
        default=0.0,
        tracking=True
    )

    # 系统 (System File and Document Admin Fee)
    outbound_system_file_charge = fields.Float(
        string='System File and Document Admin Fee',
        default=10.0,
        tracking=True
    )
    # 危险品 (Dangerous Goods Declaration Charge)
    outbound_DGD_charge = fields.Float(
        string='Dangerous Goods Declaration Charge',
        default=7.5,
        tracking=True
    )

    # 出库操作 (Outbound Handling)
    outbound_handling_charge = fields.Float(
        string='Outbound Handling',
        default=0.0,
        tracking=True
    )

    # 扫描费用
    outbound_scanning_charge = fields.Float(string='Scanning Charge', default=0.0, tracking=True)
    # 托盘
    outbound_pallet_type = fields.Selection(
        [('fba-stamp', 'FBA Stamp'), ('non-stamp', 'Non Stamp')],
        string='Pallet Type',
        default='fba-stamp'
    )
    outbound_pallet_fee = fields.Float(string='Outbound Pallet Fee', default=0.0, tracking=True)

    # 打托
    outbound_palletizing_qty = fields.Float(string='Palletizing Qty', default=1.0, tracking=True)
    outbound_palletizing_charge = fields.Float(string='Palletizing Charge', default=0.0, tracking=True)

    confirm_user_id = fields.Many2one(
        'res.users', string='Confirmed By', readonly=True, help="User who confirmed the order."
        , tracking=True)
    confirm_time_user_tz = fields.Datetime(
        string='Confirm Time (User Timezone)', readonly=True, help="Confirmation time in the user's timezone."
        , tracking=True)
    confirm_time_server = fields.Datetime(
        string='Confirm Time (Server)', readonly=True, help="Confirmation time in the server's timezone."
        , tracking=True)

    pallet_prefix_code = fields.Char(

        string="Pallet Prefix",

        index=True,  # Optimize prefix searches [4](@ref)

        help="Client-specific pallet grouping identifier (e.g., AX20250404335)"

    )
    is_auto_moves = fields.Boolean(
        string="Auto Create Move Lines",
        default=True,
        help="Automatically create stock moves for this outbound order."
    )
    # hoymiles 特有字段 delivery instruction
    delivery_issuance_time = fields.Datetime(string='Delivery Issue Time', readonly=True,
                                             help='Time when the order can be delivered')
    delivery_issuance_remark = fields.Text(string='Delivery Issue Description',
                                           help='Description of the delivery instruction')
    
    outbound_order_product_serial_numbers = fields.One2many('world.depot.outbound.order.product.serial.number','outbound_order_id')

    '''
    @api.model
    def _search_default_state_not_cancel(self):
        return [('state', '!=', 'cancel')]
    '''

    @api.model
    def create(self, values):
        """
        generate bill number
        """
        times = fields.Date.today()
        values['billno'] = self.env['ir.sequence'].next_by_code('seq.outbound.order', times)
        return super(OutboundOrder, self).create(values)

    def action_confirm(self):
        """
        Confirm the outbound order
        """

        for record in self:
            if record.state != 'new':
                raise UserError(_("Outbound order can only be confirmed from 'New' state."))
            else:
                record.state = 'confirm'
                # Record the user who confirmed
                record.confirm_user_id = self.env.user
                # Record the confirmation time in the user's timezone
                user_time = fields.Datetime.context_timestamp(self, fields.Datetime.now())
                record.confirm_time_user_tz = fields.Datetime.to_string(user_time)
                # Record the server confirmation time
                record.confirm_time_server = fields.Datetime.now()

                # Automatically add followers
                try:
                    inventory_admin_group = self.env.ref('stock.group_stock_manager')
                    inventory_user_group = self.env.ref('stock.group_stock_user')
                    valid_users = (inventory_admin_group.users | inventory_user_group.users).filtered('active')
                    partner_ids = valid_users.partner_id.ids
                    record.message_subscribe(partner_ids=partner_ids)
                    record.message_post(
                        body=Markup(
                            "Outbound Order <a href='#' data-oe-model='world.depot.outbound.order' data-oe-id='%d'>%s</a> has been confirmed, please check it."
                        ) % (record.id, record.billno),
                        partner_ids=partner_ids,
                        subtype_id=self.env.ref('mail.mt_comment').id,  # Ensures HTML rendering
                        message_type='comment',  # Explicitly set the message type
                        notify=False  # Disable email notifications
                    )

                except Exception as e:
                    _logger.error("出库单关注者通知失败: %s | 单据: %s", str(e), record.billno)
                    # 可选：raise UserError(_("通知发送失败，请手动检查"))  # 关键业务场景可中断
        return True

    def action_cancel(self):
        """ Cancel the outbound order
        """
        for record in self:
            if record.state == 'cancel':
                raise UserError(_("This order %s has already been canceled.") % record.reference)

            if record.state == 'confirm':
                if record.picking_PICK:
                    if record.picking_PICK.state == 'done':
                        raise UserError(
                            _("Cannot cancel the order %s with an active stock picking that is done.") % record.reference)
                    # If the stock picking is not done, delete it
                    try:
                        record.picking_PICK.unlink()
                    except Exception as e:
                        raise UserError(
                            _("Failed to delete stock picking for order %s: %s") % (record.reference, str(e)))

            record.state = 'cancel'

    def unlink(self):
        for record in self:
            # Check state
            if record.state not in ['new', 'cancel']:
                raise UserError(_("Only new or cancelled orders can be deleted."))

            # Iterate through all fields of the model
            for field_name, field in record._fields.items():
                if field.type == 'one2many':
                    # Get the one2many records
                    one2many_records = record[field_name]
                    one2many_records.unlink()

        return super(OutboundOrder, self).unlink()

    def action_unconfirm(self):
        """ Unconfirm the outbound order
        """
        for record in self:
            if record.state != 'confirm':
                raise UserError(_("Outbound order can only be unconfirmed from 'Confirm' state."))
            related_picking = self.env['stock.picking'].search(
                [('outbound_order_id', '=', record.id), ('state', '!=', 'cancel')], limit=1
            )
            if related_picking:
                raise UserError(_("Cannot unconfirm an order with completed stock picking."))

            record.state = 'new'
            record.confirm_user_id = False
            record.confirm_time_user_tz = False
            record.confirm_time_server = False

    def action_create_picking_PICK(self):
        """
        Create a stock picking for the outbound order
        """
        if self.state != 'confirm':
            raise UserError(_("Outbound order must be confirmed before creating a stock picking."))
        if not self.pick_type:
            raise UserError(_("Picking type must be set before creating a stock picking."))
        if not self.p_date:
            raise UserError(_("Planning date must be set before creating a stock picking."))
        if not self.reference:
            raise UserError(_("Reference must be set before creating a stock picking."))

        for record in self:
            # Check if stock picking already exists
            existing_picking = self.env['stock.picking'].search(
                [('outbound_order_id', '=', record.id),
                 ('picking_type_id', '=', record.pick_type.id),
                 ('state', '!=', 'cancel')],
                limit=1)
            if existing_picking:
                raise UserError(_("A stock picking already exists for this Outbound Order."))

            picking = self.env['stock.picking'].create({
                'picking_type_id': record.pick_type.id,
                'location_id': record.pick_type.default_location_src_id.id,
                'location_dest_id': record.pick_type.default_location_dest_id.id,
                'origin': record.billno,
                'partner_id': record.unload_company.id,
                'outbound_order_id': record.id,
                'planning_date': record.p_date,
                'ref_1': record.reference,
                'load_ref': record.load_ref,
            })

            # Create stock moves
            for product in record.outbound_order_product_ids:
                stock_move = self.env['stock.move'].create({
                    'name': product.product_id.name,
                    'product_id': product.product_id.id,
                    'product_uom_qty': product.quantity,
                    'product_uom': product.product_id.uom_id.id,
                    'picking_id': picking.id,
                    'location_id': picking.location_id.id,
                    'location_dest_id': picking.location_dest_id.id,
                    'outbound_order_product_id': product.id,
                })
                if record.is_auto_moves:
                    # Find pallets matching prefix with available stock
                    prefix = product.pallet_prefix_code or ''
                    # Build ilike pattern: when prefix is empty, match all packages; otherwise match containing the prefix
                    like_prefix = '%'
                    if prefix:
                        like_prefix = f'%-{prefix}-%'
                    pallets = self.env['stock.quant.package'].search([
                        ('name', '=ilike', like_prefix),
                        ('quant_ids.quantity', '>', 0),  # Directly filter on related quant fields
                        ('quant_ids.product_id', '=', product.product_id.id),
                        ('quant_ids.location_id.usage', '=', 'internal'), 
                        ('quant_ids.location_id.name', '!=', 'Output'), 
                        # Ensure pallets have stock for the specific product
                    ], order='create_date,name')  # Prioritize oldest pallets first
                    moves = []
                    remaining_qty = product.quantity
                    pallet_locations = []  # collect used pallet location names for this product
                    for pallet in pallets:
                        # Get total on-hand qty for the pallet
                        pallet_qty = sum(pallet.quant_ids.mapped('quantity'))
                        alloc_qty = min(pallet_qty, remaining_qty)

                        if alloc_qty <= 0:
                            continue

                        # record the location name for later writing into product.locations
                        loc_name = False
                        if pallet.location_id:
                            loc_name = pallet.location_id.complete_name
                        if loc_name:
                            # keep order and uniqueness
                            if loc_name not in pallet_locations:
                                pallet_locations.append(loc_name)

                        # Create move line for full/partial pallet
                        if product.product_id.tracking == 'serial':
                            for i_index in range(1, int(alloc_qty) + 1):
                                moves.append({
                                    'move_id': stock_move.id,
                                    'picking_id': picking.id,
                                    'product_id': product.product_id.id,
                                    'product_uom_id': product.product_id.uom_id.id,
                                    'quantity': 1,  # Serial numbers are tracked one by one
                                    'location_id': pallet.location_id.id,
                                    'location_dest_id': picking.location_dest_id.id,
                                    'package_id': pallet.id,
                                    'owner_id': pallet.owner_id.id,
                                })
                        else:
                            moves.append({
                                'move_id': stock_move.id,
                                'picking_id': picking.id,
                                'product_id': product.product_id.id,
                                'product_uom_id': product.product_id.uom_id.id,
                                'quantity': alloc_qty,  # Planned quantity
                                'location_id': pallet.location_id.id,
                                'location_dest_id': picking.location_dest_id.id,
                                'package_id': pallet.id,
                                'owner_id': pallet.owner_id.id,
                            })
                        remaining_qty -= alloc_qty

                        if remaining_qty <= 0:
                            break  # Exit when requirement met
                    # Handle insufficient stock
                    if remaining_qty > 0:
                        # Compose a helpful error message including product name and pallet prefix
                        prod_name = ''
                        try:
                            prod_name = product.product_id.name if getattr(product, 'product_id', False) else (getattr(product, 'product_name', '') or '')
                        except Exception:
                            prod_name = ''
                        raise UserError(f"Insufficient stock for {prod_name} (prefix: {prefix})! Shortfall: {remaining_qty} units")

                    # write collected pallet locations into the product.locations field
                    try:
                        if pallet_locations:
                            product.locations = ', '.join(pallet_locations)
                    except Exception:
                        # fallback: do not interrupt picking creation for writable issues
                        _logger.exception('Failed to write pallet locations for product %s', product.id)

                    self.env['stock.move.line'].create(moves)

            # Update the stock picking reference in the outbound order
            record.picking_PICK = picking.id

            # return a success message
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Stock Picking Created'),
                    'message': _('Stock picking has been created successfully.'),
                    'sticky': False,
                }
            }
            
    # New method to check stock availability without creating pickings
    def action_check_avaliable(self):
        """
        Check whether pallets exist to allocate all outbound order products.
        This performs the same allocation checks as `action_create_picking_PICK`
        but does not create pickings or move lines. It raises a UserError
        with a helpful message when allocation is impossible or partial.
        Returns True when all products can be fully allocated from matching pallets.
        """
        all_errors = []
        for record in self:
            # Only check products when auto allocation by pallets is enabled
            for product in record.outbound_order_product_ids:
                if not record.is_auto_moves:
                    continue
                prefix = product.pallet_prefix_code or ''
                # Build ilike pattern: when prefix is empty, match all packages; otherwise match containing the prefix
                like_prefix = '%'
                if prefix:
                    like_prefix = f'%-{prefix}-%'
                pallets = self.env['stock.quant.package'].search([
                    ('name', '=ilike', like_prefix),
                    ('quant_ids.quantity', '>', 0),
                    ('quant_ids.product_id', '=', product.product_id.id),
                    ('quant_ids.location_id.usage', '=', 'internal'), 
                    ('quant_ids.location_id.name', '!=', 'Output'), 
                ], order='create_date,name')

                remaining_qty = float(product.quantity or 0)
                allocated_any = False
                for pallet in pallets:
                    try:
                        pallet_qty = float(sum(pallet.quant_ids.mapped('quantity')) or 0)
                    except Exception:
                        pallet_qty = 0.0
                    alloc_qty = min(pallet_qty, remaining_qty)
                    if alloc_qty <= 0:
                        continue
                    allocated_any = True
                    remaining_qty -= alloc_qty
                    if remaining_qty <= 0:
                        break

                # build per-product error messages (do not raise immediately so we can report all)
                if not allocated_any:
                    prod_name = ''
                    try:
                        prod_name = product.product_id.name or ''
                    except Exception:
                        prod_name = ''
                    all_errors.append(f"Insufficient stock for {prod_name} (prefix: {prefix})! No allocatable pallets found.")
                elif remaining_qty > 0:
                    prod_name = ''
                    try:
                        prod_name = product.product_id.name or ''
                    except Exception:
                        prod_name = ''
                    all_errors.append(f"Insufficient stock for {prod_name} (prefix: {prefix})! Shortfall: {int(remaining_qty)} units")

        if all_errors:
            # raise a single aggregated error to show all shortages at once
            raise UserError('\n'.join(all_errors))

        # If single record call from UI, return a client notification for UX; otherwise just True
        if len(self) == 1:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Stock Availability Check Passed'),
                    'message': _('All products can be fully allocated from matching pallets.'),
                    'sticky': False,
                }
            }
        return True

    @api.depends('Outbound_order_product_ids')
    def _onchange_sum(self):
        """Update pallets field based on Outbound order products."""
        total_pallets = sum(
            product.pallets for product in self.outbound_order_product_ids if product.is_outbound_handling)
        scanning_quantity = sum(product.quantity for product in self.outbound_order_product_ids if product.is_scanning)
        self.pallets = total_pallets
        self.scanning_quantity = scanning_quantity
        self.is_adr = any(product.adr for product in self.outbound_order_product_ids)

    def action_create_cmr(self):
        for rec in self:
            # Load template
            template_data = base64.b64decode(rec.project.outbound_cmr_template_file)
            template_buffer = BytesIO(template_data)
            workbook = openpyxl.load_workbook(template_buffer, read_only=False)
            worksheet = workbook.active

            # Define alignment styles
            ALIGN_TOP_RIGHT = Alignment(vertical="top", horizontal="right", wrap_text=True)
            ALIGN_TOP_LEFT = Alignment(vertical="top", horizontal="left", wrap_text=True)
            ALIGN_TOP_CENTER = Alignment(vertical="top", horizontal="center", wrap_text=True)

            # Fill company information
            worksheet['B3'] = rec.owner.name or ''
            warehouse_address = [
                rec.warehouse.partner_id.street or '',
                rec.warehouse.partner_id.zip or '',
                rec.warehouse.partner_id.city or ''
            ]
            worksheet['B4'] = ', '.join(warehouse_address)
            worksheet['B8'] = rec.unload_company.name or ''
            unload_company_address = [
                rec.unload_company.street or '',
                rec.unload_company.zip or '',
                rec.unload_company.city or ''
            ]
            worksheet['B9'] = ', '.join(unload_company_address)
            worksheet['E19'] = rec.load_ref or ''
            worksheet['D16'] = rec.load_date.strftime('%d/%m/%Y') if rec.load_date else ''

            # Set initial row positions
            row_index = 24  # Starting row for product data
            template_row = 24  # Template row to copy styles from

            # Process each product in the outbound order
            for product in rec.outbound_order_product_ids:
                self.insert_row_manually(worksheet, row_index, template_row)
                # Fill product data with proper alignment
                # Column B: Commodity/Product Name
                worksheet.cell(row=row_index, column=2, value=product.product_id.name or '').alignment = ALIGN_TOP_LEFT
                # Column C: Quantity
                quantity = product.quantity or 0.0
                worksheet.cell(row=row_index, column=3, value=quantity).alignment = ALIGN_TOP_RIGHT
                # Column D: Net Weight per Box (NW kg/Box)
                weight_per_box = product.product_id.weight or 0.0
                worksheet.cell(row=row_index, column=4, value=weight_per_box).alignment = ALIGN_TOP_RIGHT
                # Column E: Gross Weight (GW kg) - calculated as quantity * weight per box
                weight_subtotal = quantity * weight_per_box
                worksheet.cell(row=row_index, column=5, value=weight_subtotal).alignment = ALIGN_TOP_RIGHT
                # Column F: Points (empty in the example)
                worksheet.cell(row=row_index, column=6, value='').alignment = ALIGN_TOP_RIGHT
                # Column G: Pallets (empty in the example)
                worksheet.cell(row=row_index, column=7, value='').alignment = ALIGN_TOP_RIGHT

                row_index += 1

            worksheet.delete_rows(row_index)
            # Save the workbook
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # Generate a random sequence (e.g., 001~999)
            random_seq = f"{random.randint(1, 999):03}"

            # Format the attachment name with the random sequence
            attachment_name = f"CMR_{rec.billno}_{random_seq}.xlsx"
            self.env['ir.attachment'].create({
                'name': attachment_name,
                'type': 'binary',
                'datas': base64.b64encode(output.read()),
                'res_model': self._name,
                'res_id': rec.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

            # Return success message and refresh the record
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('CMR Created'),
                    'message': _('The CMR has been successfully created and attached.'),
                    'sticky': False,
                }
            }

    def insert_row_manually(self, worksheet, new_row_index, template_row_index):
        """
        Copy styles, data, and borders from a template row to a new row.
        :param worksheet: Target worksheet object
        :param new_row_index: Index of the new row to insert
        :param template_row_index: Index of the template row
        """
        # Insert new row
        worksheet.insert_rows(new_row_index)

        # Define default border style based on your template image
        # Using thin black lines as shown in the image
        default_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Copy each cell in the template row
        for col in range(1, worksheet.max_column + 1):
            template_cell = worksheet.cell(row=template_row_index, column=col)
            new_cell = worksheet.cell(row=new_row_index, column=col)

            # 1. Copy data
            new_cell.value = template_cell.value

            # 2. Copy styles (font, fill, alignment, etc.)
            if template_cell.has_style:
                new_cell.font = copy(template_cell.font)
                new_cell.fill = copy(template_cell.fill)
                new_cell.alignment = copy(template_cell.alignment)
                new_cell.number_format = template_cell.number_format

            # 3. Apply borders - manually copy border properties to avoid recursion issues
            if template_cell.border:
                # Create new Border object with properties from template cell
                # This avoids the deepcopy recursion error while preserving border styles
                new_cell.border = Border(
                    left=template_cell.border.left,
                    right=template_cell.border.right,
                    top=template_cell.border.top,
                    bottom=template_cell.border.bottom,
                    diagonal=template_cell.border.diagonal,
                    diagonal_direction=template_cell.border.diagonal_direction,
                    outline=template_cell.border.outline,
                    vertical=template_cell.border.vertical,
                    horizontal=template_cell.border.horizontal
                )
            else:
                new_cell.border = default_border  # Apply default if no border in template

        # 4. Copy row height to maintain consistent row sizing
        if template_row_index in worksheet.row_dimensions:
            worksheet.row_dimensions[new_row_index].height = worksheet.row_dimensions[template_row_index].height

    def cron_update_outbound_date(self):
        """Scheduled action to update inbound dates for confirmed orders without an inbound date."""
        orders = self.search([])
        for order in orders:
            stock_picking = self.env['stock.picking'].search(
                [('outbound_order_id', '=', order.id), ('state', '!=', 'cancel')],
                order='scheduled_date asc',
                limit=1
            )

            if not order.delivery_street:
                order.delivery_street = order.unload_company.street or ''
            if not order.delivery_city:
                order.delivery_city = order.unload_company.city or ''
            if not order.delivery_zip:
                order.delivery_zip = order.unload_company.zip or ''
            if not order.delivery_country_id:
                order.delivery_country_id = order.unload_company.country_id or False
            if not order.delivery_phone:
                order.delivery_phone = order.unload_company.phone or ''
            if not order.delivery_mobile:
                order.delivery_mobile = order.unload_company.mobile or ''

            if stock_picking:
                # Update the stock picking ID and inbound date
                order.picking_PICK = stock_picking.id
                if stock_picking.date_done:
                    order.picking_PICK_date = stock_picking.date_done
                    if order.status == 'planning':
                        order.status = 'picking'

                    outbound = self.env['stock.picking'].search(
                        [('origin', '=', order.picking_PICK.name), ('picking_type_code', '=', 'outgoing')], limit=1)
                    if outbound:
                        order.picking_Out = outbound.id
                        if outbound.date_done:
                            order.picking_Out_date = outbound.date_done
                            order.status = 'outbound'

                    _logger.info("Updated o_date for order %s to %s", order.id, stock_picking.date_done)
            else:
                _logger.info("No valid stock picking found for order %s", order.id)

    def view_outbound_order_product_details(self):
        return {
            'name': _('Outbound Order Product Details'),
            'type': 'ir.actions.act_window',
            'res_model': 'world.depot.outbound.order.product',
            'view_mode': 'list',
            'domain': [('outbound_order_id', '=', self.id)],
            'context': {'create': False},
        }

    # View Outbound Order Serial Number Details
    @api.model
    def view_outbound_order_sn_details(self, order_id=None):
        """
        Prepare and open a list view of serial/lot details for an outbound order.
        This method accepts an optional positional order_id because the web client
        may call it with the record id as a positional argument when invoking RPC.
        """
        sn_model = self.env['world.depot.outbound.order.product.serial.number']
        try:
            # Resolve the record whether we're called on a model or with an id
            if order_id:
                rec = self.browse(order_id)
            else:
                # If called on a recordset, use the first record
                rec = self
            if not rec:
                # Return an empty list view if no record is available
                return {
                    'name': _('Outbound Order Product Details'),
                    'type': 'ir.actions.act_window',
                    'res_model': 'world.depot.outbound.order.product.serial.number',
                    'view_mode': 'list',
                    'domain': [('outbound_order_id', '=', order_id or False)],
                    'context': {'create': False},
                }

            # Prepare data for bulk insertion
            sn_details = []

            picking = rec.picking_PICK
            if not picking:
                # No picking associated; return an empty list view
                return {
                    'name': _('Outbound Order Product Details'),
                    'type': 'ir.actions.act_window',
                    'res_model': 'world.depot.outbound.order.product.serial.number',
                    'view_mode': 'list',
                    'domain': [('outbound_order_id', '=', rec.id)],
                    'context': {'create': False},
                }

            if getattr(picking, 'state', '') == 'done':
                # Iterate move lines directly to avoid extra searches
                moves=self.env['stock.move'].search([('picking_id', '=', picking.id)])
                for move in moves:
                    move_lines = self.env['stock.move.line'].search([('move_id', '=', move.id)])
                    for move_line in move_lines:
                        if not move_line.product_id:
                            _logger.warning("Product missing for move line ID %s in order ID %s", move_line.id, self.id)
                            continue
                        sn_details.append({
                            'outbound_order_id': rec.id,
                            'type': rec.type or '',
                            'reference': rec.reference or '',
                            'p_date': picking.date_done or False,
                            'project': rec.project.id or False,
                            'picking_PICK': picking.id or False,
                            'product_id': move_line.product_id.id or False,
                            'product_name': move_line.product_id.name or '',
                            'lot_id': move_line.lot_id.id or False,
                            'lot_name': move_line.lot_id.name or '',
                            'quantity': move_line.quantity or 0,
                        })

            # Perform delete/create in a savepoint so failures don't leave data half-updated
            with self.env.cr.savepoint():
                existing = sn_model.search([('outbound_order_id', '=', rec.id)])
                if existing:
                    existing.unlink()
                if sn_details:
                    sn_model.create(sn_details)

            return {
                'name': _('Outbound Order Product Details'),
                'type': 'ir.actions.act_window',
                'res_model': 'world.depot.outbound.order.product.serial.number',
                'view_mode': 'list',
                'domain': [('outbound_order_id', '=', rec.id)],
                'context': {'create': False},
            }

        except Exception as e:
            _logger.exception('Error initializing OutboundOrderSNDetail for order %s: %s', getattr(rec, 'id', order_id), e)
            raise UserError(_('Failed to initialize SN details: %s') % e)
        

class OutboundOrderProduct(models.Model):
    _name = 'world.depot.outbound.order.product'
    _description = 'Outbound Order Product'

    outbound_order_id = fields.Many2one('world.depot.outbound.order', string='Outbound Order', required=True)
    project = fields.Many2one(related='outbound_order_id.project', string='Project', store=True, readonly=True)
    project_category_id = fields.Many2one(related='project.category', string='Project Category', store=True,
                                          readonly=True)
    cntr_no = fields.Char(string='Container No', required=False)
    product_id = fields.Many2one('product.product', string='Product', required=True, tracking=True,
                                 domain="[('categ_id', '=', project_category_id)]")
    adr = fields.Boolean(string='ADR', help='Indicates if the product is classified as ADR (dangerous goods)',
                         related='product_id.is_dg', stock=True)
    un_number = fields.Char(string='UN Number', help='United Nations number for dangerous goods classification',
                            related='product_id.un_code', store=True)
    pallets = fields.Float(string='Pallets', required=True)
    pallet_type = fields.Char(string='Pallet Type', help='How many quantity on a pallet', default='')
    pallet_no = fields.Char(string='Pallet No', help='Pallet number for tracking', default='')
    quantity = fields.Float(string='Quantity', default=1.0, required=True, tracking=True, )
    remark = fields.Text(string='Remark')
    is_serial_tracked = fields.Boolean(string='Tracked by Serial', compute='_compute_is_serial_tracked', store=True)
    serial_numbers = fields.Text(string='Serial Numbers',
                                 help="Comma-separated list of serial numbers for the product.")
    is_outbound_handling = fields.Boolean(string='is Handling', default=True, tracking=True)
    outbound_handling_price = fields.Float(string='Handling Price', default=True, tracking=True)
    outbound_handling_unit = fields.Selection(
        string='Unit',
        selection=[
            ('pallet', 'Per Pallet'),
            ('piece', 'Per Piece')],
        readonly=True,
    )
    outbound_handling_charge = fields.Float(string='Handling Charge', default=0.0, tracking=True)
    is_scanning = fields.Boolean(string='is Scanning', default=True, tracking=True)
    outbound_scanning_price = fields.Float(string='Scanning Price', default=0.0, tracking=True)
    outbound_scanning_charge = fields.Float(string='Scanning Charge', default=0.0, tracking=True)
    #product_serial_number_ids = fields.One2many('world.depot.outbound.order.product.serial.number',
    #                                            'inbound_order_product_id', string='Product Serial Numbers')
    pallet_prefix_code = fields.Char(

        string="Pallet Prefix",

        index=True,  # Optimize prefix searches [4](@ref)

        help="Client-specific pallet grouping identifier (e.g., AX20250404335)"

    )

    barcode = fields.Char(string='Barcode', related='product_id.barcode', store=True, readonly=True)
    default_code = fields.Char(string='Default Code', related='product_id.default_code', store=True, readonly=True)
    weight = fields.Float(string='Weight', related='product_id.weight', store=True, readonly=True)
    weight_subtotal = fields.Float(string='Weight Subtotal', compute='_compute_weight_subtotal', store=True)
    
    locations=fields.Char(string='Locations', help='Storage locations of the product in the warehouse')
    
    outbound_order_product_serial_numbers = fields.One2many('world.depot.outbound.order.product.serial.number','outbound_order_product_id')

    @api.depends('product_id')
    def _compute_is_serial_tracked(self):
        for record in self:
            record.is_serial_tracked = record.product_id.tracking == 'serial'
            '''
            record.is_scanning = record.product_id.tracking == 'serial' or record.product_id.tracking == 'lot'
            record.outbound_handling_price = record.outbound_order_id.project.outbound_handling_price
            record.outbound_handling_unit = record.outbound_order_id.project.outbound_handling_unit
            record.outbound_scanning_price = record.outbound_order_id.project.outbound_scanning_price
            '''

    @api.depends('pallets', 'quantity', 'outbound_handling_price')
    def _compute_outbound_handling_charge(self):
        """Compute the outbound handling charge based on pallets, quantity, and project settings."""
        for record in self:
            record.outbound_handling_charge = 0
            record.outbound_handling_price = 0
            '''
            if record.is_outbound_handling:
                record.outbound_handling_price = record.outbound_order_id.project.outbound_handling_price
                record.outbound_handling_unit = record.outbound_order_id.project.outbound_handling_per
                if record.outbound_order_id.project.outbound_handling_per == 'pallet':
                    record.outbound_handling_charge = record.pallets * record.outbound_order_id.project.outbound_handling_price
                elif record.outbound_order_id.project.outbound_handling_per == 'piece':
                    record.outbound_handling_charge = record.quantity * record.outbound_order_id.project.outbound_handling_price
                else:
                    record.outbound_handling_unit = False
                    record.outbound_handling_charge = 0
                    record.outbound_handling_price = 0
            '''

    @api.depends('is_scanning', 'quantity', 'outbound_scanning_price')
    def _compute_outbound_scanning_charge(self):
        """Compute the Outbound scanning charge based on quantity and project settings."""
        for record in self:
            record.outbound_scanning_charge = 0
            record.outbound_scanning_price = 0
            '''
            if record.is_scanning:
                record.outbound_scanning_price = record.outbound_order_id.project.outbound_scanning_price
                record.outbound_scanning_charge = record.quantity * record.outbound_order_id.project.outbound_scanning_price
            else:
                record.outbound_scanning_price = 0.0
            '''

    @api.depends('product_id', 'quantity', 'product_id.weight')
    def _compute_weight_subtotal(self):
        for record in self:
            record.weight_subtotal = (record.product_id.weight or 0.0) * (record.quantity or 0.0)



# 其他附件
class OutboundOderDocs(models.Model):
    _name = 'world.depot.outbound.order.docs'
    _description = 'world.depot.outbound.order.docs'

    doc_type = fields.Selection(
        selection=[
            ('cmr', 'CMR'),
            ('sn_details', 'SN Details'),
            ('other', 'Other Document'),
        ],
        string="Document Type",
        required=True,
        tracking=True
    )
    description = fields.Text(string='Description')
    file = fields.Binary(string='File')
    filename = fields.Char(string='File name')
    outbound_order_id = fields.Many2one('world.depot.outbound.order', string='Outbound Order', required=True)


class OutboundOrderProductSerialNumber(models.Model):
    _description = 'Outbound Order Product Serial Number'
    _name = 'world.depot.outbound.order.product.serial.number'

    # Support linking either to an outbound order product record (for
    # per-product serial details) or directly to an outbound order (the
    # aggregated SN detail listing). Both fields are optional so the same
    # model can be used by both flows.
    outbound_order_product_id = fields.Many2one('world.depot.outbound.order.product', string='Outbound Order Product')
    outbound_order_id = fields.Many2one('world.depot.outbound.order', string='Outbound Order', readonly=True)
    type = fields.Char(string='Type', readonly=True)
    reference = fields.Char(string='Outbound Reference', readonly=True)
    p_date = fields.Date(string='Date', readonly=True)
    project = fields.Many2one('project.project', string='Project', readonly=True)
    project_name = fields.Char(string='Project Name', readonly=True, related='project.name')
    picking_PICK = fields.Many2one('stock.picking', string='Picking', readonly=True)
    product_id = fields.Many2one('product.product', string='Product')
    product_name = fields.Char(string='Product Name', readonly=True)
    lot_id = fields.Many2one('stock.lot', string='Serial/Lot ID', readonly=True)
    lot_name = fields.Char(string='Serial/Lot Name', readonly=True)
    quantity = fields.Float(string='Quantity', readonly=True)

    
