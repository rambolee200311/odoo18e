# -*- coding: utf-8 -*-

import base64
import os
from io import BytesIO

from odoo import _, fields, models
from odoo.exceptions import UserError
try:
    import openpyxl
except ImportError:
    openpyxl = None
import logging

_logger = logging.getLogger(__name__)


class SunriseProductMasterImport(models.Model):
    _name = "stock.barcode.lite.sunrise.product.master.import"
    _description = "Sunrise Product Master Import"
    _order = "id desc"

    name = fields.Char(string="Name", required=True, default="Sunrise Product Master Import", copy=False, index=True)
    file = fields.Binary(string="Excel File", required=True, copy=False)
    filename = fields.Char(string="Filename", copy=False)
    state = fields.Selection([("draft", "Draft"), ("done", "Done"), ("partial", "Partial"), ("failed", "Failed")], string="State", default="draft", required=True, copy=False, index=True)
    remark = fields.Text(string="Remark", copy=False)
    success_count = fields.Integer(string="Success Count", copy=False)
    failed_count = fields.Integer(string="Failed Count", copy=False)
    sheet_name = fields.Char(string="Sheet Name", copy=False)
    line_ids = fields.One2many("stock.barcode.lite.sunrise.product.master.import.line", "import_id", string="Import Lines", copy=False)
    message = fields.Text(string="Import Message", copy=False)

    def action_import_sunrise_product_master(self):
        for rec in self:
            if rec.state != "draft":
                rec.write({"remark": _("Only draft import records can be imported.")})
                continue

            _logger.info(
                "Sunrise product master import %s started. file=%s sheet=%s",
                rec.name,
                rec.filename,
                rec.sheet_name or "ALL",
            )
            rec.write({"message": _("Reading Excel file...")})

            try:
                category = rec.get_sunrise_product_category()
                rows = rec.read_sunrise_product_master_rows()
            except Exception as error:
                _logger.exception("Sunrise product master import %s failed while reading Excel.", rec.name)
                rec.write({
                    "state": "failed",
                    "remark": str(error),
                    "message": _("Import failed while reading Excel: %s") % str(error),
                    "success_count": 0,
                    "failed_count": 0,
                })
                continue

            _logger.info(
                "Sunrise product master import %s read %s rows.",
                rec.name,
                len(rows),
            )
            rec.write({"message": _("Read %s Excel rows. Preparing import lines...") % len(rows)})

            seen_codes = {}
            for row in rows:
                line_values_list = rec.prepare_sunrise_product_import_line_values(row, category)
                for line_values in line_values_list:
                    product_code = line_values.get("product_code")
                    _logger.info(
                        "Sunrise product master import %s processing code=%s sheet=%s row=%s",
                        rec.name,
                        product_code,
                        line_values.get("sheet_name"),
                        line_values.get("row_number"),
                    )
                    rec.write({
                        "message": _("Processing %(code)s at %(sheet)s row %(row)s...") % {
                            "code": product_code,
                            "sheet": line_values.get("sheet_name"),
                            "row": line_values.get("row_number"),
                        }
                    })
                    if line_values.get("remark"):
                        _logger.info(
                            "Sunrise product master import %s skipped code=%s before create. reason=%s",
                            rec.name,
                            product_code,
                            line_values.get("remark"),
                        )
                        rec.line_ids.create(line_values)
                        continue
                    if product_code in seen_codes:
                        first_line = seen_codes[product_code]
                        duplicate_remark = _(
                            "Product code is duplicated in this Excel file.\n"
                            "Current: sheet %(current_sheet)s row %(current_row)s (%(current_code_type)s).\n"
                            "First found: sheet %(first_sheet)s row %(first_row)s (%(first_code_type)s)."
                        ) % {
                            "current_sheet": line_values.get("sheet_name"),
                            "current_row": line_values.get("row_number"),
                            "current_code_type": line_values.get("code_type"),
                            "first_sheet": first_line.get("sheet_name"),
                            "first_row": first_line.get("row_number"),
                            "first_code_type": first_line.get("code_type"),
                        }
                        is_shared_semi_finished_suspected = (
                            line_values.get("code_type") == "semi_finished"
                            and first_line.get("code_type") == "semi_finished"
                            and line_values.get("product_name")
                            and line_values.get("product_name") == first_line.get("product_name")
                        )
                        if (
                            is_shared_semi_finished_suspected
                            and first_line.get("counterpart_code")
                            and line_values.get("counterpart_code")
                        ):
                            duplicate_remark = "%s\n%s" % (
                                duplicate_remark,
                                _("Current finished code: %(current_finished_code)s. First finished code: %(first_finished_code)s.") % {
                                    "current_finished_code": line_values.get("counterpart_code"),
                                    "first_finished_code": first_line.get("counterpart_code"),
                                },
                            )
                        line_values.update({
                            "state": "failed",
                            "remark": duplicate_remark,
                            "is_shared_semi_finished_suspected": is_shared_semi_finished_suspected,
                        })
                        _logger.info(
                            "Sunrise product master import %s duplicate code in Excel: %s first_sheet=%s first_row=%s first_type=%s",
                            rec.name,
                            product_code,
                            first_line.get("sheet_name"),
                            first_line.get("row_number"),
                            first_line.get("code_type"),
                        )
                        rec.line_ids.create(line_values)
                        continue
                    seen_codes[product_code] = {
                        "sheet_name": line_values.get("sheet_name"),
                        "row_number": line_values.get("row_number"),
                        "code_type": line_values.get("code_type"),
                        "product_name": line_values.get("product_name"),
                        "counterpart_code": line_values.get("counterpart_code"),
                    }

                    existing_product = rec.env["product.product"].sudo().search([("barcode", "=", product_code)], limit=1)
                    if existing_product:
                        line_values.update({
                            "state": "failed",
                            "remark": _("Product code already exists."),
                            "product_id": existing_product.id,
                        })
                        _logger.info(
                            "Sunrise product master import %s existing product code=%s product_id=%s",
                            rec.name,
                            product_code,
                            existing_product.id,
                        )
                        rec.line_ids.create(line_values)
                        continue

                    line = rec.line_ids.create(line_values)
                    try:
                        with rec.env.cr.savepoint():
                            product = rec.create_sunrise_product_from_import_line(line)
                    except Exception as error:
                        _logger.exception(
                            "Sunrise product master import %s failed code=%s sheet=%s row=%s",
                            rec.name,
                            product_code,
                            line_values.get("sheet_name"),
                            line_values.get("row_number"),
                        )
                        line.write({
                            "state": "failed",
                            "remark": str(error),
                        })
                    else:
                        _logger.info(
                            "Sunrise product master import %s success code=%s product_id=%s",
                            rec.name,
                            product_code,
                            product.id,
                        )
                        line.write({
                            "state": "success",
                            "remark": _("Imported successfully."),
                            "product_id": product.id,
                        })

            rec.update_sunrise_product_import_result()

        next_action = False
        if self:
            next_action = {
                "type": "ir.actions.act_window",
                "res_model": "stock.barcode.lite.sunrise.product.master.import",
                "view_mode": "form",
                "views": [(False, "form")],
                "res_id": self[:1].id,
                "target": "current",
            }

        params = {
            "title": _("Sunrise Product Master Import"),
            "message": _("Import finished."),
            "sticky": False,
        }
        if next_action:
            params["next"] = next_action

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": params,
        }

    def get_sunrise_product_category(self):
        self.ensure_one()
        category_model = self.env["product.category"]
        category = category_model.sudo().search([("complete_name", "=", "All / SUNRISE")], limit=1)
        if not category:
            raise ValueError(_('Product category "All / SUNRISE" was not found.'))
        return category_model.browse(category.id)

    def read_sunrise_product_master_rows(self):
        self.ensure_one()
        if not self.file:
            raise ValueError(_("Please upload an Excel file."))

        extension = os.path.splitext(self.filename or "")[1].lower()
        if extension != ".xlsx":
            raise ValueError(_("Only .xlsx files are supported."))
        if openpyxl is None:
            raise ValueError(_("openpyxl is required to import .xlsx files."))

        file_content = base64.b64decode(self.file)
        workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        rows = []
        target_sheet_name = self.cell_to_text(self.sheet_name)
        core_fields = ["finished_code", "product_name", "english_name", "aux_uom_name", "tax_name"]

        for sheet in workbook.worksheets:
            sheet_name = self.cell_to_text(sheet.title)
            if sheet_name == "特殊标记说明":
                continue
            if target_sheet_name and sheet_name != target_sheet_name:
                continue

            _logger.info("Sunrise product master import %s reading sheet: %s", self.name, sheet_name)
            self.write({"message": _("Reading sheet: %s") % sheet_name})

            header_map = {}
            header_row_number = 0
            for row_number, row_values in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                candidate_map = self.get_sunrise_header_map(row_values)
                if all(field_name in candidate_map for field_name in ("finished_code", "product_name")):
                    header_map = candidate_map
                    header_row_number = row_number
                    break

            if not header_map:
                raise ValueError(_('Sheet "%s" has no valid header row.') % sheet_name)

            missing_headers = [field_name for field_name in core_fields if field_name not in header_map]
            if missing_headers:
                raise ValueError(
                    _('Sheet "%s" is missing required headers: %s') % (sheet_name, ", ".join(missing_headers)))

            for row_number, row_values in enumerate(sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
                                                    start=header_row_number + 1):
                if not any(self.cell_to_text(value) for value in row_values):
                    continue
                rows.append({
                    "sheet_name": sheet_name,
                    "row_number": row_number,
                    "finished_code": self.get_sunrise_cell_value(row_values, header_map, "finished_code"),
                    "semi_finished_code": self.get_sunrise_cell_value(row_values, header_map, "semi_finished_code"),
                    "product_name": self.get_sunrise_cell_value(row_values, header_map, "product_name"),
                    "english_name": self.get_sunrise_cell_value(row_values, header_map, "english_name"),
                    "chinese_spec": self.get_sunrise_cell_value(row_values, header_map, "chinese_spec"),
                    "english_spec": self.get_sunrise_cell_value(row_values, header_map, "english_spec"),
                    "tax_name": self.get_sunrise_cell_value(row_values, header_map, "tax_name"),
                    "main_uom_name": self.get_sunrise_cell_value(row_values, header_map, "main_uom_name"),
                    "aux_uom_name": self.get_sunrise_cell_value(row_values, header_map, "aux_uom_name"),
                    "uom_conversion_rate": self.get_sunrise_cell_value(row_values, header_map, "uom_conversion_rate"),
                    "shelf_life_months": self.get_sunrise_cell_value(row_values, header_map, "shelf_life_months"),
                    "shelf_life_years": self.get_sunrise_cell_value(row_values, header_map, "shelf_life_years"),
                })

        if not rows:
            if target_sheet_name:
                raise ValueError(_('Sheet "%s" has no data rows or was not found.') % target_sheet_name)
            raise ValueError(_("The Excel file has no data rows."))
        return rows

    def get_sunrise_header_map(self, row_values):
        self.ensure_one()
        aliases = {
            "finished_code": ["成品编码"],
            "semi_finished_code": ["半成品编码"],
            "product_name": ["中文产品名称"],
            "english_name": ["英文产品名称"],
            "chinese_spec": ["中文规格"],
            "english_spec": ["英文规格"],
            "tax_name": ["税目"],
            "main_uom_name": ["主计量单位"],
            "aux_uom_name": ["辅计量单位"],
            "uom_conversion_rate": ["与主计量单位换算系数", "换算系数"],
            "shelf_life_months": ["保质期/月"],
            "shelf_life_years": ["保质期/年"],
        }
        header_map = {}
        for index, value in enumerate(row_values):
            header_text = self.normalize_sunrise_header(value)
            if not header_text:
                continue
            for field_name, field_aliases in aliases.items():
                if field_name in header_map:
                    continue
                for alias in field_aliases:
                    alias_text = self.normalize_sunrise_header(alias)
                    if alias_text and alias_text in header_text:
                        header_map[field_name] = index
                        break
        return header_map

    def normalize_sunrise_header(self, value):
        return self.cell_to_text(value).replace("\n", "").replace("\r", "").replace(" ", "").replace("　", "").lower()

    def get_sunrise_cell_value(self, row_values, header_map, field_name):
        self.ensure_one()
        if field_name not in header_map:
            return ""
        index = header_map[field_name]
        if index >= len(row_values):
            return ""
        return self.cell_to_text(row_values[index])

    def prepare_sunrise_product_import_line_values(self, row, category):
        self.ensure_one()
        product_name = row.get("product_name", "")
        english_name = row.get("english_name", "")
        chinese_spec = row.get("chinese_spec", "")
        english_spec = row.get("english_spec", "")
        description_parts = []
        if chinese_spec:
            description_parts.append(_("Chinese Spec: %s") % chinese_spec)
        if english_spec:
            description_parts.append(_("English Spec: %s") % english_spec)
        description = "\n".join(description_parts)
        parse_errors = []
        uom_conversion_rate = 0.0
        shelf_life_months = 0
        shelf_life_years = 0
        try:
            uom_conversion_rate = self.cell_to_float(row.get("uom_conversion_rate"))
        except ValueError:
            parse_errors.append(_("UOM conversion rate must be a number."))
        try:
            shelf_life_months = self.cell_to_integer(row.get("shelf_life_months"))
        except ValueError:
            parse_errors.append(_("Shelf life months must be a number."))
        try:
            shelf_life_years = self.cell_to_integer(row.get("shelf_life_years"))
        except ValueError:
            parse_errors.append(_("Shelf life years must be a number."))

        base_values = {
            "import_id": self.id,
            "sheet_name": row.get("sheet_name", ""),
            "row_number": row.get("row_number", 0),
            "product_name": product_name,
            "english_name": english_name,
            "chinese_spec": chinese_spec,
            "english_spec": english_spec,
            "description": description,
            "category_id": category.id,
            "tracking": "lot",
            "attribute_name": "包装规格",
            "attribute_value_name": "标准包装",
            "main_uom_name": row.get("main_uom_name", ""),
            "aux_uom_name": row.get("aux_uom_name", ""),
            "uom_conversion_rate": uom_conversion_rate,
            "shelf_life_months": shelf_life_months,
            "shelf_life_years": shelf_life_years,
            "tax_name": row.get("tax_name", ""),
            "state": "failed",
            "remark": "; ".join(parse_errors),
        }

        line_values_list = []
        finished_code = row.get("finished_code", "")
        semi_finished_code = row.get("semi_finished_code", "")
        if finished_code == "/":
            finished_code = ""
        if semi_finished_code == "/":
            semi_finished_code = ""
        if finished_code:
            values = dict(base_values)
            values.update({
                "code_type": "finished",
                "product_code": finished_code,
                "counterpart_code": semi_finished_code,
            })
            line_values_list.append(values)
        # if semi_finished_code:
        #     values = dict(base_values)
        #     values.update({
        #         "code_type": "semi_finished",
        #         "product_code": semi_finished_code,
        #         "counterpart_code": finished_code,
        #     })
        #     line_values_list.append(values)
        return line_values_list

    def create_sunrise_product_from_import_line(self, line):
        self.ensure_one()
        if not line.product_code:
            raise ValueError(_("Product code is required."))
        if not line.product_name and not line.english_name:
            raise ValueError(_("Chinese product name or English product name is required."))
        if not line.aux_uom_name:
            raise ValueError(_("Auxiliary UOM is required."))
        if not line.tax_name:
            raise ValueError(_("Tax item is required."))

        uom = self.get_or_create_sunrise_uom(line.aux_uom_name)
        tax = self.get_or_create_sunrise_sale_tax(line.tax_name)
        attribute, attribute_value = self.get_or_create_sunrise_standard_package_attribute()
        template_model = self.env["product.template"]

        template = template_model.create({
            "name": line.english_name or line.product_name,
            "categ_id": line.category_id.id,
            "tracking": "lot",
            "type": "consu",
            "is_storable": True,
            "description": line.description,
            "uom_id": uom.id,
            "uom_po_id": uom.id,
            "taxes_id": [(6, 0, [tax.id])],
            "sunrise_chinese_name": line.product_name,
            "sunrise_english_name": line.english_name,
            "sunrise_main_uom_name": line.main_uom_name,
            "sunrise_uom_conversion_rate": line.uom_conversion_rate,
            "sunrise_shelf_life_months": line.shelf_life_months,
            "sunrise_shelf_life_years": line.shelf_life_years,
            "attribute_line_ids": [(0, 0, {
                "attribute_id": attribute.id,
                "value_ids": [(6, 0, [attribute_value.id])],
            })],
        })

        products = template.product_variant_ids
        if len(products) != 1:
            raise ValueError(_("The standard package variant was not created correctly."))

        product = self.env["product.product"].browse(products.id)
        product.write({
            "default_code": line.product_code,
            "barcode": line.product_code,
        })
        line.write({
            "uom_id": uom.id,
            "uom_po_id": uom.id,
            "tax_id": tax.id,
        })
        return product

    def get_or_create_sunrise_uom(self, uom_name):
        self.ensure_one()
        uom_model = self.env["uom.uom"]
        category_model = self.env["uom.category"]
        uom = uom_model.sudo().search([("name", "=", uom_name)], limit=1)
        if uom:
            return uom_model.browse(uom.id)

        category = category_model.create({"name": "Sunrise %s" % uom_name})
        return uom_model.create({
            "name": uom_name,
            "category_id": category.id,
            "uom_type": "reference",
            "rounding": 0.01,
        })

    def get_or_create_sunrise_sale_tax(self, tax_name):
        self.ensure_one()
        tax_model = self.env["account.tax"]
        tax = tax_model.sudo().search([
            ("name", "=", tax_name),
            ("type_tax_use", "=", "sale"),
            ("company_id", "in", [False, self.env.company.id]),
        ], limit=1)
        if tax:
            return tax_model.browse(tax.id)
        return tax_model.create({
            "name": tax_name,
            "type_tax_use": "sale",
            "amount_type": "percent",
            "amount": 0.0,
            "company_id": self.env.company.id,
        })

    def get_or_create_sunrise_standard_package_attribute(self):
        self.ensure_one()
        attribute_model = self.env["product.attribute"]
        value_model = self.env["product.attribute.value"]
        attribute = attribute_model.sudo().search([("name", "=", "包装规格")], limit=1)
        if attribute:
            attribute = attribute_model.browse(attribute.id)
        else:
            attribute = attribute_model.create({
                "name": "包装规格",
                "create_variant": "always",
                "display_type": "radio",
            })

        attribute_value = value_model.sudo().search([
            ("attribute_id", "=", attribute.id),
            ("name", "=", "标准包装"),
        ], limit=1)
        if attribute_value:
            attribute_value = value_model.browse(attribute_value.id)
        else:
            attribute_value = value_model.create({
                "attribute_id": attribute.id,
                "name": "标准包装",
            })
        return attribute, attribute_value

    def update_sunrise_product_import_result(self):
        for rec in self:
            success_count = len(rec.line_ids.filtered(lambda line: line.state == "success"))
            failed_count = len(rec.line_ids.filtered(lambda line: line.state == "failed"))
            if success_count and failed_count:
                state = "partial"
            elif success_count:
                state = "done"
            else:
                state = "failed"
            rec.write({
                "state": state,
                "success_count": success_count,
                "failed_count": failed_count,
                "remark": _("Success: %(success)s, Failed: %(failed)s") % {
                    "success": success_count,
                    "failed": failed_count,
                },
                "message": _("Import finished. Success: %(success)s, Failed: %(failed)s") % {
                    "success": success_count,
                    "failed": failed_count,
                },
            })
            _logger.info(
                "Sunrise product master import %s finished. success=%s failed=%s state=%s",
                rec.name,
                success_count,
                failed_count,
                state,
            )

    def cell_to_text(self, value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def cell_to_float(self, value):
        value_text = self.cell_to_text(value)
        if not value_text:
            return 0.0
        return float(value_text)

    def cell_to_integer(self, value):
        value_text = self.cell_to_text(value)
        if not value_text:
            return 0
        return int(float(value_text))


class SunriseProductMasterImportLine(models.Model):
    _name = "stock.barcode.lite.sunrise.product.master.import.line"
    _description = "Sunrise Product Master Import Line"
    _order = "id desc"

    import_id = fields.Many2one("stock.barcode.lite.sunrise.product.master.import", string="Import", required=True, ondelete="cascade", copy=False, index=True)
    sheet_name = fields.Char(string="Sheet Name", copy=False, index=True)
    row_number = fields.Integer(string="Row Number", copy=False)
    code_type = fields.Selection([("finished", "Finished"), ("semi_finished", "Semi Finished")], string="Code Type", copy=False, index=True)
    product_code = fields.Char(string="Product Code", copy=False, index=True)
    counterpart_code = fields.Char(string="Counterpart Code", copy=False, index=True)
    product_name = fields.Char(string="Chinese Product Name", copy=False)
    english_name = fields.Char(string="English Name", copy=False)
    chinese_spec = fields.Char(string="Chinese Spec", copy=False)
    english_spec = fields.Char(string="English Spec", copy=False)
    description = fields.Text(string="Internal Description", copy=False)
    category_id = fields.Many2one("product.category", string="Category", copy=False, index=True)
    tracking = fields.Selection([("lot", "By Lots")], string="Tracking", copy=False)
    attribute_name = fields.Char(string="Attribute Name", copy=False)
    attribute_value_name = fields.Char(string="Attribute Value Name", copy=False)
    main_uom_name = fields.Char(string="Main UOM Name", copy=False)
    aux_uom_name = fields.Char(string="Auxiliary UOM Name", copy=False)
    uom_id = fields.Many2one("uom.uom", string="Unit of Measure", copy=False, index=True)
    uom_po_id = fields.Many2one("uom.uom", string="Purchase Unit of Measure", copy=False, index=True)
    uom_conversion_rate = fields.Float(string="UOM Conversion Rate", copy=False)
    shelf_life_months = fields.Integer(string="Shelf Life Months", copy=False)
    shelf_life_years = fields.Integer(string="Shelf Life Years", copy=False)
    tax_name = fields.Char(string="Tax Name", copy=False)
    tax_id = fields.Many2one("account.tax", string="Sales Tax", copy=False, index=True)
    product_id = fields.Many2one("product.product", string="Product", copy=False, index=True)
    is_shared_semi_finished_suspected = fields.Boolean(string="Suspected Shared Semi Finished", default=False, copy=False, index=True)
    state = fields.Selection([("success", "Success"), ("failed", "Failed")], string="State", default="failed", copy=False, index=True)
    remark = fields.Text(string="Remark", copy=False)
    is_retry_success = fields.Boolean(string="Retry Success", default=False, copy=False, index=True)
    def action_retry_import(self):
        success_count = 0
        failed_count = 0
        import_records = self.mapped("import_id")
        product_model = self.env["product.product"]

        for rec in self:
            if rec.state != "failed":
                raise UserError(_("Only failed import lines can be retried."))
            if not rec.import_id:
                raise UserError(_("The import record is required."))

            rec.import_id.write({
                "message": _("Retrying product code %s...") % rec.product_code,
            })

            existing_product = product_model.sudo().with_context(active_test=False).search([
                ("barcode", "=", rec.product_code),
            ], limit=1)

            if existing_product:
                new_remark = _('Retry failed: product code "%s" already exists.') % rec.product_code
                rec.write({
                    "product_id": existing_product.id,
                    "remark": "%s\n%s" % (rec.remark, new_remark) if rec.remark else new_remark,
                })
                failed_count += 1
                continue

            try:
                with self.env.cr.savepoint():
                    product = rec.import_id.create_sunrise_product_from_import_line(rec)
            except Exception as error:
                _logger.exception(
                    "Retry Sunrise product import failed. line=%s code=%s",
                    rec.id,
                    rec.product_code,
                )
                new_remark = _("Retry failed: %s") % str(error)
                rec.write({
                    "state": "failed",
                    "remark": "%s\n%s" % (rec.remark, new_remark) if rec.remark else new_remark,
                })
                failed_count += 1
            else:
                new_remark = _("Imported successfully after retry.")
                rec.write({
                    "state": "success",
                    "product_id": product.id,
                    "is_shared_semi_finished_suspected": False,
                    "remark": "%s\n%s" % (rec.remark, new_remark) if rec.remark else new_remark,
                    "is_retry_success": True,
                })
                success_count += 1

        import_records.update_sunrise_product_import_result()

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Retry Product Import"),
                "message": _("Retry finished. Success: %(success)s, Failed: %(failed)s") % {
                    "success": success_count,
                    "failed": failed_count,
                },
                "type": "success" if not failed_count else "warning",
                "sticky": False,
                "next": {
                    "type": "ir.actions.client",
                    "tag": "reload",
                },
            },
        }