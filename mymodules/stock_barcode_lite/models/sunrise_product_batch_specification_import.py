# -*- coding: utf-8 -*-

import base64
import logging
import math
import os
import re
from io import BytesIO

from odoo import _, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None
_logger = logging.getLogger(__name__)

class SunriseProductBatchSpecificationImport(models.Model):
    _name = "sunrise.product.specification.import"
    _description = "Sunrise Product Batch Specification Import"
    _order = "id desc"

    name = fields.Char(string="Name", required=True, default="Sunrise Product Batch Specification Import", copy=False, index=True)
    file = fields.Binary(string="Excel File", required=True, copy=False)
    filename = fields.Char(string="Filename", copy=False)
    state = fields.Selection([("draft", "Draft"), ("done", "Done"), ("partial", "Partial"), ("failed", "Failed")], string="State", default="draft", required=True, copy=False, index=True)
    total_count = fields.Integer(string="Total Count", copy=False)
    success_count = fields.Integer(string="Success Count", copy=False)
    warning_count = fields.Integer(string="Warning Count", copy=False)
    failed_count = fields.Integer(string="Failed Count", copy=False)
    remark = fields.Text(string="Remark", copy=False)
    message = fields.Text(string="Import Message", copy=False)
    line_ids = fields.One2many("sunrise.product.specification.import.line", "import_id", string="Import Lines", copy=False)

    def action_import_sunrise_product_batch_specification(self):
        for rec in self:
            if rec.state != "draft":
                rec.write({"remark": _("Only draft import records can be imported.")})
                continue

            try:
                rows = rec.read_sunrise_product_batch_specification_rows()
            except Exception as error:
                _logger.exception("Sunrise product batch specification import %s failed while reading Excel.", rec.name)
                rec.write({"state": "failed", "remark": str(error), "message": str(error), "total_count": 0, "success_count": 0, "warning_count": 0, "failed_count": 0})
                continue

            seen_batch_rows = {}
            success_count = 0
            warning_count = 0
            failed_count = 0
            for row in rows:
                line = rec.line_ids.create(rec.prepare_sunrise_product_batch_specification_line_values(row))
                try:
                    with rec.env.cr.savepoint():
                        line_state = rec.process_sunrise_product_batch_specification_line(line, seen_batch_rows)
                except Exception as error:
                    failed_count += 1
                    line.write({"state": "failed", "error_msg": str(error)})
                    continue

                if line_state == "warning":
                    warning_count += 1
                else:
                    success_count += 1

            rec.write({
                "total_count": len(rows),
                "success_count": success_count,
                "warning_count": warning_count,
                "failed_count": failed_count,
                "state": rec.get_sunrise_product_batch_specification_import_state(success_count, warning_count, failed_count),
                "message": _("Import finished. Success: %(success)s, Warnings: %(warning)s, Failed: %(failed)s.") % {
                    "success": success_count,
                    "warning": warning_count,
                    "failed": failed_count,
                },
            })
        return True

    def update_sunrise_product_batch_specification_import_result(self):
        for rec in self:
            success_count = len(rec.line_ids.filtered(lambda line: line.state == "success"))
            warning_count = len(rec.line_ids.filtered(lambda line: line.state == "warning"))
            failed_count = len(rec.line_ids.filtered(lambda line: line.state == "failed"))
            rec.write({
                "total_count": len(rec.line_ids),
                "success_count": success_count,
                "warning_count": warning_count,
                "failed_count": failed_count,
                "state": rec.get_sunrise_product_batch_specification_import_state(success_count, warning_count, failed_count),
            })

    def get_sunrise_product_batch_specification_import_state(self, success_count, warning_count, failed_count):
        if failed_count and not success_count and not warning_count:
            return "failed"
        if failed_count or warning_count:
            return "partial"
        return "done"

    def get_sunrise_existing_batch_row_signatures(self):
        self.ensure_one()
        signatures = {}
        for line in self.line_ids.filtered(lambda item: item.state in ("success", "warning") and item.product_id and item.lot_id):
            signatures[(line.product_id.id, line.lot_id.id)] = line.get_sunrise_product_batch_specification_signature()
        return signatures

    def process_sunrise_product_batch_specification_line(self, line, seen_batch_rows):
        self.ensure_one()
        if line.error_msg:
            raise UserError(line.error_msg)
        product_model = self.env["product.product"]
        product_template_model = self.env["product.template"]
        lot_model = self.env["stock.lot"]
        product_ids = product_model.sudo().search([("barcode", "=", line.product_code)]).ids
        if not product_ids:
            raise UserError(_("Row %(row)s: product barcode %(code)s was not found.") % {"row": line.row_number, "code": line.product_code})
        if len(product_ids) > 1:
            raise UserError(_("Row %(row)s: product barcode %(code)s matches multiple products.") % {"row": line.row_number, "code": line.product_code})

        product = product_model.browse(product_ids[0])
        lot_ids = lot_model.sudo().search([("product_id", "=", product.id), ("name", "=", line.lot_name)]).ids
        if not lot_ids:
            raise UserError(_("Row %(row)s: batch %(lot)s for product %(code)s was not found. The import does not create batches.") % {"row": line.row_number, "lot": line.lot_name, "code": line.product_code})
        if len(lot_ids) > 1:
            raise UserError(_("Row %(row)s: batch %(lot)s matches multiple records for product %(code)s.") % {"row": line.row_number, "lot": line.lot_name, "code": line.product_code})

        lot = lot_model.browse(lot_ids[0])
        line.write({"product_id": product.id, "product_template_id": product.product_tmpl_id.id, "lot_id": lot.id})
        batch_key = (product.id, lot.id)
        signature = line.get_sunrise_product_batch_specification_signature()
        previous_signature = seen_batch_rows.get(batch_key)
        if previous_signature and previous_signature != signature:
            raise UserError(_("Row %(row)s: the same product and batch has different values in this Excel file.") % {"row": line.row_number})
        seen_batch_rows[batch_key] = signature

        product_template = product_template_model.browse(product.product_tmpl_id.id)
        product_dimensions = self.convert_sunrise_product_dimensions_to_meters(line.product_dimensions, line.row_number)
        warnings = []
        template_values = {}
        if line.gross_weight:
            if not product_template.gross_weight:
                template_values["gross_weight"] = line.gross_weight
            elif not math.isclose(product_template.gross_weight, line.gross_weight, rel_tol=1e-9, abs_tol=1e-6):
                warnings.append(_("Product template gross weight %(existing)s differs from Excel %(incoming)s; template was not changed.") % {"existing": product_template.gross_weight, "incoming": line.gross_weight})
        if product_dimensions:
            if not product_template.product_dimensions:
                template_values["product_dimensions"] = product_dimensions
            elif self.normalize_sunrise_product_dimensions(product_template.product_dimensions) != self.normalize_sunrise_product_dimensions(product_dimensions):
                warnings.append(_("Product template dimensions differ from Excel; template was not changed."))
        if line.sunrise_uom_conversion_rate:
            if not product_template.sunrise_uom_conversion_rate:
                template_values["sunrise_uom_conversion_rate"] = line.sunrise_uom_conversion_rate
            elif not math.isclose(product_template.sunrise_uom_conversion_rate, line.sunrise_uom_conversion_rate, rel_tol=1e-9, abs_tol=1e-6):
                warnings.append(_("Product template conversion rate %(existing)s differs from Excel %(incoming)s; template was not changed.") % {"existing": product_template.sunrise_uom_conversion_rate, "incoming": line.sunrise_uom_conversion_rate})
        if template_values:
            product_template.write(template_values)

        lot_values = {}
        if line.gross_weight and not math.isclose(lot.gross_weight, line.gross_weight, rel_tol=1e-9, abs_tol=1e-6):
            lot_values["gross_weight"] = line.gross_weight
        if product_dimensions and self.normalize_sunrise_product_dimensions(lot.product_dimensions) != self.normalize_sunrise_product_dimensions(product_dimensions):
            lot_values["product_dimensions"] = product_dimensions
        if lot_values:
            lot.write(lot_values)

        line.write({"state": "warning" if warnings else "success", "remark": "\n".join(warnings), "error_msg": False})
        return "warning" if warnings else "success"

    def prepare_sunrise_product_batch_specification_line_values(self, row):
        self.ensure_one()
        errors = []
        gross_weight = self.parse_sunrise_product_batch_specification_float(row.get("gross_weight"), "gross_weight", row["row_number"], errors)
        conversion_rate = self.parse_sunrise_product_batch_specification_float(row.get("sunrise_uom_conversion_rate"), "sunrise_uom_conversion_rate", row["row_number"], errors)
        product_code = self.get_sunrise_text(row.get("product_code"))
        lot_name = self.get_sunrise_text(row.get("lot_name"))
        if not product_code:
            errors.append(_("Row %s: product code is required.") % row["row_number"])
        if not lot_name:
            errors.append(_("Row %s: batch number is required.") % row["row_number"])
        return {
            "import_id": self.id,
            "sheet_name": row["sheet_name"],
            "row_number": row["row_number"],
            "product_code": product_code,
            "lot_name": lot_name,
            "gross_weight": gross_weight,
            "product_dimensions": self.get_sunrise_text(row.get("product_dimensions")),
            "sunrise_uom_conversion_rate": conversion_rate,
            "state": "failed",
            "error_msg": "\n".join(errors),
        }

    def read_sunrise_product_batch_specification_rows(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Please upload an Excel file."))
        if openpyxl is None:
            raise UserError(_("openpyxl is required to import .xlsx files."))
        if os.path.splitext(self.filename or "")[1].lower() != ".xlsx":
            raise UserError(_("Only .xlsx files are supported."))

        workbook = openpyxl.load_workbook(BytesIO(base64.b64decode(self.file)), read_only=True, data_only=True)
        rows = []
        try:
            for sheet in workbook.worksheets:
                if sheet.title == "特殊标记说明":
                    continue
                header_map = {}
                header_row_number = 0
                for row_number, row_values in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                    header_map = self.get_sunrise_product_batch_specification_header_map(row_values)
                    if all(field_name in header_map for field_name in ("product_code", "gross_weight", "product_dimensions", "sunrise_uom_conversion_rate", "lot_name")):
                        header_row_number = row_number
                        break
                if not header_row_number:
                    raise UserError(_("Sheet %s is missing required headers.") % sheet.title)
                for row_number, row_values in enumerate(sheet.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
                    if not any(self.get_sunrise_text(value) for value in row_values):
                        continue
                    rows.append({
                        "sheet_name": sheet.title,
                        "row_number": row_number,
                        "product_code": self.get_sunrise_cell_value(row_values, header_map, "product_code"),
                        "gross_weight": self.get_sunrise_cell_value(row_values, header_map, "gross_weight"),
                        "product_dimensions": self.get_sunrise_cell_value(row_values, header_map, "product_dimensions"),
                        "sunrise_uom_conversion_rate": self.get_sunrise_cell_value(row_values, header_map, "sunrise_uom_conversion_rate"),
                        "lot_name": self.get_sunrise_cell_value(row_values, header_map, "lot_name"),
                    })
        finally:
            workbook.close()
        if not rows:
            raise UserError(_("The Excel file has no data rows."))
        return rows

    def get_sunrise_product_batch_specification_header_map(self, row_values):
        self.ensure_one()
        aliases = {
            "product_code": ["存货编码", "product_code", "barcode"],
            "gross_weight": ["毛重/kg", "毛重", "gross_weight"],
            "product_dimensions": ["箱规/cm", "箱规", "产品尺寸", "product_dimensions"],
            "sunrise_uom_conversion_rate": ["千粒数/箱", "换算率", "sunrise_uom_conversion_rate"],
            "lot_name": ["批次号", "lot_name"],
        }
        header_map = {}
        for index, value in enumerate(row_values):
            header_text = self.normalize_sunrise_header(value)
            for field_name, field_aliases in aliases.items():
                if field_name in header_map:
                    continue
                if any(self.normalize_sunrise_header(alias) in header_text for alias in field_aliases):
                    header_map[field_name] = index
        return header_map

    def get_sunrise_cell_value(self, row_values, header_map, field_name):
        index = header_map.get(field_name)
        if index is None or index >= len(row_values):
            return ""
        return self.get_sunrise_text(row_values[index])

    def parse_sunrise_product_batch_specification_float(self, value, field_name, row_number, errors):
        value_text = self.get_sunrise_text(value)
        if not value_text:
            return 0.0
        try:
            number = float(value_text.replace(",", ""))
        except (TypeError, ValueError):
            errors.append(_("Row %(row)s: %(field)s must be a positive number.") % {"row": row_number, "field": field_name})
            return 0.0
        if not math.isfinite(number) or number <= 0:
            errors.append(_("Row %(row)s: %(field)s must be a positive number.") % {"row": row_number, "field": field_name})
            return 0.0
        return number

    def normalize_sunrise_product_dimensions(self, value):
        return "".join(self.get_sunrise_text(value).upper().split())

    def convert_sunrise_product_dimensions_to_meters(self, value, row_number):
        raw_value = self.get_sunrise_text(value)
        if not raw_value:
            return ""

        normalized_value = re.sub(r"\s*(cm|厘米)\s*", "", raw_value, flags=re.IGNORECASE)
        parts = re.split(r"\s*[*×xX]\s*", normalized_value)
        if len(parts) != 3:
            raise UserError(_("Row %(row)s: product dimensions must be length*width*height in cm.") % {"row": row_number})

        try:
            dimensions = [float(part) / 100 for part in parts]
        except (TypeError, ValueError) as error:
            raise UserError(_("Row %(row)s: product dimensions must contain valid numbers.") % {"row": row_number}) from error

        if any(not math.isfinite(dimension) or dimension <= 0 for dimension in dimensions):
            raise UserError(_("Row %(row)s: product dimensions must be positive numbers.") % {"row": row_number})

        return "*".join(f"{dimension:.6f}".rstrip("0").rstrip(".") for dimension in dimensions)

    def normalize_sunrise_header(self, value):
        return self.get_sunrise_text(value).replace("\n", "").replace("\r", "").replace(" ", "").replace("　", "").lower()

    def get_sunrise_text(self, value):
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()


class SunriseProductBatchSpecificationImportLine(models.Model):
    _name = "sunrise.product.specification.import.line"
    _description = "Sunrise Product Batch Specification Import Line"
    _order = "id desc"

    import_id = fields.Many2one("sunrise.product.specification.import", string="Import", required=True, ondelete="cascade", copy=False, index=True)
    sheet_name = fields.Char(string="Sheet Name", copy=False, index=True)
    row_number = fields.Integer(string="Row Number", copy=False, index=True)
    product_code = fields.Char(string="Product Code", copy=False, index=True)
    lot_name = fields.Char(string="Batch Number", copy=False, index=True)
    product_id = fields.Many2one("product.product", string="Product", copy=False, index=True)
    product_template_id = fields.Many2one("product.template", string="Product Template", copy=False, index=True)
    lot_id = fields.Many2one("stock.lot", string="Batch", copy=False, index=True)
    gross_weight = fields.Float(string="Gross Weight (kg)", copy=False)
    product_dimensions = fields.Char(string="Product Dimensions (cm)", copy=False)
    sunrise_uom_conversion_rate = fields.Float(string="Standard Box In Qty", copy=False)
    state = fields.Selection([("success", "Success"), ("warning", "Warning"), ("failed", "Failed")], string="State", default="failed", required=True, copy=False, index=True)
    error_msg = fields.Text(string="Error", copy=False)
    remark = fields.Text(string="Remark", copy=False)

    def get_sunrise_product_batch_specification_signature(self):
        self.ensure_one()
        return (
            round(self.gross_weight, 6),
            "".join((self.product_dimensions or "").upper().split()),
            round(self.sunrise_uom_conversion_rate, 6),
        )

    def action_retry_import(self):
        for rec in self:
            if rec.state != "failed":
                raise UserError(_("Only failed lines can be retried."))
            import_record = rec.import_id
            if not import_record:
                raise UserError(_("The import record is missing."))

            seen_batch_rows = import_record.get_sunrise_existing_batch_row_signatures()
            rec.write({"state": "failed", "error_msg": False})
            try:
                with import_record.env.cr.savepoint():
                    import_record.process_sunrise_product_batch_specification_line(rec, seen_batch_rows)
            except Exception as error:
                rec.write({"state": "failed", "error_msg": str(error)})
            import_record.update_sunrise_product_batch_specification_import_result()
        return True
